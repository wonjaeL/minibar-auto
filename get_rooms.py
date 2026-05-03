import sys
import openpyxl

FILE = "7.Daily Minibar Sold_2026_04.xlsx"


def get_room_list(day: int, file_path: str = FILE) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = str(day)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    rooms = []
    for col in range(8, ws.max_column + 1):
        v = ws.cell(row=5, column=col).value
        if v is None or v == "":
            continue
        s = str(int(v)) if isinstance(v, (int, float)) else str(v).strip()
        if not s.isdigit():
            continue
        if len(s) == 3:
            s = "0" + s
        rooms.append(s)
    return ",".join(rooms)


if __name__ == "__main__":
    day = int(sys.argv[1]) if len(sys.argv) > 1 else 28
    print(get_room_list(day))
